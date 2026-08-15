from typing import List, Optional
import io
import openpyxl
from decimal import Decimal
from sqlalchemy.ext.asyncio import AsyncSession
from backend.app.domains.catalog.repositories import ProductRepository, CategoryRepository, PriceBookRepository
from backend.app.domains.catalog.models import Product, Category, PriceBook, PriceBookEntry
from backend.app.domains.catalog.schemas import (
    ProductCreate, ProductUpdate,
    CategoryCreate, CategoryUpdate,
    PriceBookCreate, PriceBookUpdate, PriceBookEntryCreate,
    ProductImportResponse, ProductImportErrorDetail,
    BillingType
)
from backend.app.core.exceptions import EntityNotFoundError, DomainValidationError

class CatalogService:
    """
    Business service layer managing the product catalog, category categorization,
    and multiple price book definitions.
    """
    def __init__(self, db: AsyncSession):
        self.db = db
        self.product_repo = ProductRepository(db)
        self.category_repo = CategoryRepository(db)
        self.price_book_repo = PriceBookRepository(db)

    async def create_product(self, schema: ProductCreate) -> Product:
        """
        Create a new product. SKU must be unique.
        """
        existing = await self.product_repo.get_by_sku(schema.sku)
        if existing:
            raise DomainValidationError(f"Product SKU '{schema.sku}' already exists.")
        
        if schema.category_id:
            category = await self.category_repo.get_by_id(schema.category_id)
            if not category:
                raise DomainValidationError(f"Category with ID {schema.category_id} does not exist.")
                
        return await self.product_repo.create(schema)

    async def get_product(self, product_id: int) -> Product:
        product = await self.product_repo.get_by_id(product_id)
        if not product:
            raise EntityNotFoundError(f"Product with ID {product_id} not found.")
        return product

    async def list_products(self, limit: int = 100, offset: int = 0, category_id: Optional[int] = None) -> List[Product]:
        return await self.product_repo.list(limit, offset, category_id)

    async def create_category(self, schema: CategoryCreate) -> Category:
        existing = await self.category_repo.get_by_name(schema.name)
        if existing:
            raise DomainValidationError(f"Category '{schema.name}' already exists.")
        return await self.category_repo.create(schema)

    async def list_categories(self) -> List[Category]:
        return await self.category_repo.list()

    async def create_price_book(self, schema: PriceBookCreate) -> PriceBook:
        """
        Create a new price book. If marked standard, disable other standard price books.
        """
        if schema.is_standard:
            existing_std = await self.price_book_repo.get_standard_price_book()
            if existing_std:
                # Toggle off the previous standard
                existing_std.is_standard = False
                self.db.add(existing_std)
                
        return await self.price_book_repo.create(schema)

    async def add_price_book_entry(self, price_book_id: int, schema: PriceBookEntryCreate) -> PriceBookEntry:
        # Verify both price book and product exist
        price_book = await self.price_book_repo.get_by_id(price_book_id)
        if not price_book:
            raise EntityNotFoundError(f"Price Book with ID {price_book_id} not found.")
            
        await self.get_product(schema.product_id)
        return await self.price_book_repo.add_entry(price_book_id, schema)

    async def get_product_price(self, product_id: int, price_book_id: Optional[int] = None) -> float:
        """
        Resolve the current unit price of a product.
        Checks custom price book first, falling back to standard price book,
        and finally the base product price.
        """
        product = await self.get_product(product_id)
        
        # 1. Check custom price book if specified
        if price_book_id:
            pb = await self.price_book_repo.get_by_id(price_book_id)
            if pb:
                for entry in pb.entries:
                    if entry.product_id == product_id and entry.is_active:
                        return entry.custom_price

        # 2. Check standard price book
        std_pb = await self.price_book_repo.get_standard_price_book()
        if std_pb:
            for entry in std_pb.entries:
                if entry.product_id == product_id and entry.is_active:
                    return entry.custom_price

        # 3. Fallback to base product list price
        return product.base_price

    async def update_product(self, product_id: int, schema: ProductUpdate) -> Product:
        product = await self.get_product(product_id)
        if schema.sku and schema.sku != product.sku:
            existing = await self.product_repo.get_by_sku(schema.sku)
            if existing:
                raise DomainValidationError(f"Product SKU '{schema.sku}' already exists.")
        
        if schema.category_id:
            category = await self.category_repo.get_by_id(schema.category_id)
            if not category:
                raise DomainValidationError(f"Category with ID {schema.category_id} does not exist.")
                
        return await self.product_repo.update(product, schema)

    async def deactivate_product(self, product_id: int) -> Product:
        product = await self.get_product(product_id)
        product.is_active = False
        self.db.add(product)
        await self.db.flush()
        return product

    async def activate_product(self, product_id: int) -> Product:
        product = await self.get_product(product_id)
        product.is_active = True
        self.db.add(product)
        await self.db.flush()
        return product

    async def import_products_from_excel(self, contents: bytes, filename: str, has_cost_manage_permission: bool = False) -> ProductImportResponse:
        # Validate file extension
        if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
            raise DomainValidationError("Invalid file extension. Only .xlsx and .xls are supported.")
            
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        except Exception as e:
            raise DomainValidationError(f"Invalid Excel file format: {str(e)}")

        sheet = workbook.active
        if not sheet or sheet.max_row < 2:
            return ProductImportResponse(total_rows=0, imported_count=0, failed_count=0, errors=[])

        # Get header columns
        headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in sheet[1]]
        
        # cost_price is optional, but if present requires permission
        cost_price_col_present = "cost_price" in headers
        if cost_price_col_present and not has_cost_manage_permission:
            raise DomainValidationError("You do not have permission to import cost information.")

        required_cols = ["name", "sku", "description", "category", "product_type", "base_price", "currency", "status"]
        for col in required_cols:
            if col not in headers:
                raise DomainValidationError(f"Missing required column header: '{col}'")

        col_indices = {col: headers.index(col) for col in headers if col}
        
        total_rows = 0
        imported_count = 0
        failed_count = 0
        errors = []

        # Cache categories list to avoid N+1 queries
        all_categories = await self.category_repo.list()
        cat_by_name = {c.name.lower(): c for c in all_categories}

        # Keep track of SKUs created in the current batch to avoid internal duplicates
        seen_skus = set()

        for row_idx in range(2, sheet.max_row + 1):
            row_cells = list(sheet[row_idx])
            
            # Check if row is entirely empty
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row_cells):
                continue
                
            total_rows += 1

            # Helper to get cell value by column name
            def get_val(col_name: str) -> Optional[str]:
                if col_name not in col_indices:
                    return None
                idx = col_indices[col_name]
                if idx < len(row_cells):
                    val = row_cells[idx].value
                    return str(val).strip() if val is not None else None
                return None

            name = get_val("name")
            sku = get_val("sku")
            description = get_val("description")
            category_name = get_val("category")
            product_type = get_val("product_type")
            base_price_str = get_val("base_price")
            currency = get_val("currency")
            status_val = get_val("status")
            crm_product_code = get_val("crm_product_code")
            billing_type = get_val("billing_type")

            # Row Validation
            row_error = None
            
            if not sku:
                row_error = "Missing SKU"
            elif not name:
                row_error = "Missing product name"
            elif sku.lower() in seen_skus:
                row_error = f"Duplicate SKU '{sku}' in import file"
            else:
                # Check DB for SKU
                existing_prod = await self.product_repo.get_by_sku(sku)
                if existing_prod:
                    row_error = "SKU already exists"

            if not row_error:
                # Validate Product Type
                if not product_type or product_type.lower() not in ["product", "bundle", "service"]:
                    row_error = "Invalid product type"
                    
            if not row_error:
                # Validate Category
                if not category_name or category_name.lower() not in cat_by_name:
                    row_error = "Invalid category"

            if not row_error:
                # Validate Price
                try:
                    price_val = Decimal(base_price_str) if base_price_str else Decimal("0")
                    if price_val < 0:
                        row_error = "Invalid price"
                except (ValueError, TypeError, ArithmeticError):
                    row_error = "Invalid price"

            if not row_error and cost_price_col_present:
                # Validate Cost
                cost_val_str = get_val("cost_price")
                if cost_val_str:
                    try:
                        cost_val = Decimal(cost_val_str)
                        if cost_val < 0:
                            row_error = "Invalid cost"
                    except (ValueError, TypeError, ArithmeticError):
                        row_error = "Invalid cost"

            if not row_error:
                # Validate Currency
                if not currency or currency.upper() != "USD":
                    row_error = "Invalid currency"

            if not row_error:
                # Validate Status
                if not status_val or status_val.lower() not in ["active", "inactive"]:
                    row_error = "Invalid status"

            if not row_error and "billing_type" in col_indices:
                # Validate Billing Type
                if not billing_type or billing_type.upper() not in ["MRC", "NRC", "USAGE"]:
                    row_error = "Invalid billing type"

            if row_error:
                errors.append(ProductImportErrorDetail(row=row_idx, sku=sku, error=row_error))
                failed_count += 1
            else:
                # Insert Product
                seen_skus.add(sku.lower())
                cat_obj = cat_by_name[category_name.lower()]
                
                cost_price_val = None
                if cost_price_col_present:
                    cost_val_str = get_val("cost_price")
                    if cost_val_str:
                        cost_price_val = Decimal(cost_val_str)

                prod_create = ProductCreate(
                    sku=sku,
                    name=name,
                    description=description,
                    base_price=Decimal(base_price_str) if base_price_str else Decimal("0"),
                    cost_price=cost_price_val,
                    currency=currency.upper() if currency else "USD",
                    is_active=(status_val.lower() == "active"),
                    category_id=cat_obj.id,
                    external_crm_id=crm_product_code,
                    billing_type=BillingType(billing_type.upper()) if billing_type else BillingType.MRC
                )
                
                await self.product_repo.create(prod_create)
                imported_count += 1

        # Commit transaction if there are successfully imported products
        if imported_count > 0:
            await self.db.commit()

        return ProductImportResponse(
            total_rows=total_rows,
            imported_count=imported_count,
            failed_count=failed_count,
            errors=errors
        )

ClassSymbol = CatalogService
